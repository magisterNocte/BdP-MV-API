"""
Überprüfung der @pfadfinden.de Adressen
Das Programm überprüft für die @pfadfinden.de Adressen, ob die Mitglieder noch aktive Funktionen auf Bundesebene haben
"""
import pathlib
from operator import itemgetter

from APITools import Console as c
from APITools import Nami
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from userconf import user

# variables
path = str(pathlib.Path(__file__).parent.resolve())
config = []
nami = Nami(config)
nami.auth(user.username, user.password)

# excel initialization
sourceWbPostfach = load_workbook(
    path + "\data\outlook_data_filtered_postfach.xlsx")
sourceWbWeiterleitung = load_workbook(
    path + "\data\outlook_data_filtered_weiterleitung.xlsx")
sourceWsPostfach = sourceWbPostfach.active
sourceWsWeiterleitung = sourceWbWeiterleitung.active


bundesTaetigkeitIDs = [
    """
    Liste aller Bundestätigkeiten IDs in der MV
    """
    "(100)",
    "(107)",
    "(108)",
    "(109)",
    "(111)",
    "(112)",
    "(114)",
    "(116)",
    "(138)",
    "(139)",
    "(140)",
    "(141)",
    "(142)",
    "(152)",
    "(163)",
    "(165)",
    "(166)",
    "(174)",
    "(175)",
    "(184)",
    "(187)",
    "(10,028)",
    "(10,035)",
    "(10,036)",
    "(10,037)",
    "(10,038)",
    "(10,039)",
    "(10,046)",
    "(10,047)",
    "(10,051)",
    "(10,053)",
    "(10,577)",
    "(10,651)",
    "(10,663)"
]


def filterTaetigkeit(userId, filterIDs):
    try:
        bundesTaetigkeitenUser = []
        for i in nami.taetigkeit(userId):
            for x in bundesTaetigkeitIDs:
                if x in i['entries_taetigkeit']:
                    if i['entries_aktivBis'] == '':
                        return([[i['entries_taetigkeit'], i['entries_aktivBis']]])
                    bundesTaetigkeitenUser.append(
                        [i['entries_taetigkeit'], str(i['entries_aktivBis'])[0:10]])

        bundesTaetigkeitenUser.sort(key=itemgetter(1))
        if bundesTaetigkeitenUser == []:
            return([["keine Bundestätigkeit", "0000-00-00"]])
        return([bundesTaetigkeitenUser[-1]])
    except:
        return([["ERROR3"], ["ERROR3"]])


def getUserBundesTätigkeit(Mitglied_Vorname, Mitglied_Nachname):
    letzteTaetigkeit = []
    letzteTaetigkeitDatum = []
    try:
        user = nami.user(Mitglied_Vorname, Mitglied_Nachname)
        if user == []:
            return([["ERROR1"], ["ERROR1"]])

        for i in user:
            if "@pfadfinden.de" in i['entries_email']:
                userId = i['entries_id']
                taetigkeit = filterTaetigkeit(userId, bundesTaetigkeitIDs)
                letzteTaetigkeit.append(taetigkeit[0][0])
                letzteTaetigkeitDatum.append(taetigkeit[0][1])
                return([letzteTaetigkeit, letzteTaetigkeitDatum])
        for i in user:
            userId = i['entries_id']
            taetigkeit = filterTaetigkeit(userId, bundesTaetigkeitIDs)
            letzteTaetigkeit.append(taetigkeit[0][0])
            letzteTaetigkeitDatum.append(taetigkeit[0][1])
        return([letzteTaetigkeit, letzteTaetigkeitDatum])

    except:
        return([["ERROR2"], ["ERROR2"]])


def getDocLength(sourceWs):
    i = 1
    while sourceWs["A" + str(i)].value != None:
        i += 1
    return(i)


def userLoop(sourceWs):
    '''
    die Funktion looped durch alle user in einer Tabelle und ergänzt die letzte Tätigkeit auf Bundesebene und das Datum, wann sie beendet wurde
        - kein Datum bedeutet dass die Tätigkeit noch nicht beendet ist
        - ERROR1 bedeutet, dass der user nicht in der MV ist
        - "keine Bundestätigkeit" und das Datum "0000-00-00" bedeutet dass der user keine Bundestätigkeit in der MV stehen hat
    '''
    i = 2
    while sourceWs["A" + str(i)].value != None:
        tätigkeitData = getUserBundesTätigkeit(
            sourceWs["A" + str(i)].value, sourceWs["B" + str(i)].value)
        column = 0
        for x in tätigkeitData[0]:
            sourceWs[get_column_letter(column*2 + 4) +
                     str(i)] = tätigkeitData[0][column]
            sourceWs[get_column_letter(column*2 + 5) +
                     str(i)] = tätigkeitData[1][column]
            column += 1

        c.printProgressBar(i + 1, getDocLength(sourceWs), prefix='Progress:',
                           suffix=sourceWs["A" + str(i)].value + " " + sourceWs["B" + str(i)].value + " " + str(sourceWs["D" + str(i)].value) + "                             ", length=75)

        i += 1


print("Überprüfe Tätigkeiten der User mit Postfächern:")
userLoop(sourceWsPostfach)
print("Überprüfe Tätigkeiten der User mit Weiterleitungen:")
userLoop(sourceWsWeiterleitung)


# save excelfiles
sourceWbPostfach.save(path + "\data\outlook_data_filtered_postfach.xlsx")
sourceWbWeiterleitung.save(
    path + "\data\outlook_data_filtered_weiterleitung.xlsx")
