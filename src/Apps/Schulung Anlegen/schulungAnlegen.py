
import pathlib

from decouple import config
from openpyxl import Workbook, load_workbook
from Tools.APITools import Nami
from Tools.EditUser import EditUser
from Tools.UserInfo import UserInfo

# variables
path = str(pathlib.Path(__file__).parent.resolve())
username = config("MVUSERNAME")
password = config("PASSWORD")
# Excel init
sourceWb = load_workbook(path + "\data\sourceData.xlsx")
sourceWs = sourceWb.active

newWb = Workbook()
newWs = newWb.active

# Nami
config = []
nami = Nami(config)
nami.auth(username, password)


def copy_line(i):
    newWs["A" + str(i)] = sourceWs["A" + str(i)].value
    newWs["B" + str(i)] = sourceWs["B" + str(i)].value
    newWs["C" + str(i)] = sourceWs["C" + str(i)].value
    newWs["D" + str(i)] = sourceWs["D" + str(i)].value
    newWs["E" + str(i)] = sourceWs["E" + str(i)].value


for i in range(2, sourceWs.max_row + 1):
    copy_line(i)
    if not sourceWs["A" + str(i)].value or not sourceWs["B" + str(i)].value:
        sourceWs["F" + str(i)] = "ERROR: Fehlerhafter Name"
        continue

    vornameTemp = sourceWs["A" + str(i)].value.strip().split()[0].split("-")[0]
    nachnameTemp = sourceWs["B" + str(i)].value.strip()

    tempUser = EditUser(
        nami, vornameTemp, nachnameTemp)

    if not isinstance(tempUser.mglied, int):
        newWs["F" + str(i)] = tempUser.mglied
        print(sourceWs["A" + str(i)].value,
              sourceWs["B" + str(i)].value, tempUser.mglied)
    else:
        tempUser.schulungAnlegen(
            21, str(sourceWs["E" + str(i)].value)[0:10], "Bula22 Intakt Schulung")
        newWs["F" +
              str(i)] = f'success {UserInfo.userSchulung(nami, tempUser.mglied, "Bula22 Intakt Schulung")}'
        print(sourceWs["A" + str(i)].value,
              sourceWs["B" + str(i)].value, "success")


sourceWb.save(path + "\\data\\sourceData.xlsx")
newWb.save(path + "\\data\\newWB.xlsx")
