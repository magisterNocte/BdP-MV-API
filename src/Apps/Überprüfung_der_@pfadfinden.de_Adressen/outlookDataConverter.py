"""
Der code erstellt aus der Outlook user.csv 
    1. eine übersichtbare Tabelle aller Daten
    2. eine Tabelle mit allen Weiterleitungen
    3. eine Tabelle mit allen Postfächern
    4. eine Tabelle mit allen Adressen die nicht nach Personen benannt wurden

Dafür muss die user.csv als outlook_data.xlsx im data order des Projekts abgespeichert werden

"""

from openpyxl import Workbook, load_workbook

from Tools.outlookDataConverterTools import excelHandling as exH


def main():
    outlookSourceWb = load_workbook(exH.path + '\..\data\outlook_data.xlsx')
    outlookSourceWs = outlookSourceWb.active
    # new Workbooks
    outlookSortedWb = Workbook()
    outlookFilteredWbPostfach = Workbook()
    outlookFilteredWbWeiterleitung = Workbook()
    outlookNotPeopleWb = Workbook()
    outlookSortedWs = outlookSortedWb.active
    outlookFilteredWsPostfach = outlookFilteredWbPostfach.active
    outlookFilteredWsWeiterleitung = outlookFilteredWbWeiterleitung.active
    outlookNotPeopleWs = outlookNotPeopleWb.active

    exH.sort_outlook_ws(outlookSourceWs, outlookSortedWs)
    exH.filter_outlook_ws(
        outlookSortedWs, outlookFilteredWsPostfach, outlookFilteredWsWeiterleitung, outlookNotPeopleWs)
    outlookSortedWb.save(exH.path + "\..\data\outlook_data_sorted.xlsx")
    outlookFilteredWbPostfach.save(
        exH.path + "\..\data\outlook_data_filtered_postfach.xlsx")
    outlookFilteredWbWeiterleitung.save(
        exH.path + "\..\data\outlook_data_filtered_weiterleitung.xlsx")
    outlookNotPeopleWb.save(exH.path + "\..\data\outlook_data_no_people.xlsx")
    outlookNotPeopleWb.save(exH.path + "\..\data\outlook_data_no_people.xlsx")


if __name__ == "__main__":
    main()
