import logging
import pathlib

from decouple import config
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from Tools.APITools import Nami
from Tools.UserInfo import UserInfo
from Tools.Utility import Utility

logging.basicConfig(level=logging.ERROR)
# variables
path = str(pathlib.Path(__file__).parent.resolve())
username = config("MVUSERNAME")
password = config("PASSWORD")


config = []
nami = Nami(config)
nami.auth(username, password)

# excel init
sourceWb = load_workbook(path + "\data\sourceDataID.xlsx")
sourceWs = sourceWb.active

newWb = Workbook()
newWs = newWb.active


def fillInExcel(userDetails, rowNum):
    if not isinstance(userDetails[0], int):
        newWs["A" + rowNum] = userDetails[0]
        newWs["B" + rowNum] = sourceWs["B" + rowNum].value
        return
    for i in range(1, sourceWs.max_column+1):
        newWs[get_column_letter(
            i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
        if sourceWs[get_column_letter(i) + "1"].value is None:
            continue
        match sourceWs[get_column_letter(i) + "1"].value.lower():
            case "vorname":
                newWs[get_column_letter(
                    i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
            case "nachname":
                newWs[get_column_letter(
                    i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
            case "name":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["vorname"] + " " + userDetails[1]["nachname"]
            case "id":
                newWs[get_column_letter(i) + rowNum] = userDetails[0]
            case "straße":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["strasse"]
            case "plz":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["plz"]
            case "plz/ort":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["plz"] + " " + userDetails[1]["ort"]
            case "ort":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["ort"]
            case "adresse":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["strasse"] + \
                    " " + userDetails[1]["plz"] + " " + userDetails[1]["ort"]
            case "geschlecht m/w":
                newWs[get_column_letter(
                    i) + rowNum] = "m" if userDetails[1]["geschlecht"] == "männlich" else "w"
            case "geschlecht":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["geschlecht"]
            case "bundesland":
                newWs[get_column_letter(
                    i) + rowNum] = Utility.plzZuBundesland(userDetails[1]["plz"])
            case "telefon":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["telefon1"] if userDetails[1]["telefon1"] != "" else userDetails[1]["telefon2"] if userDetails[1]["telefon2"] != "" else userDetails[1]["telefon3"]
            case "telefonnummer":
                newWs[get_column_letter(
                    i) + rowNum] = userDetails[1]["telefon1"] if userDetails[1]["telefon1"] != "" else userDetails[1]["telefon2"] if userDetails[1]["telefon2"] != "" else userDetails[1]["telefon3"]
            case "lv":
                newWs[get_column_letter(
                    i) + rowNum] = Utility.stammesIdToLV(userDetails[1]["gruppierung"])[1]
            case "geburtstag":
                newWs[get_column_letter(
                    i) + rowNum] = Utility.formatDate(str(userDetails[1]["geburtsDatum"])[0:10])
            case "geburtsdatum":
                newWs[get_column_letter(
                    i) + rowNum] = Utility.formatDate(str(userDetails[1]["geburtsDatum"])[0:10])
            case "email":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["email"]
            case "efz":
                newWs[get_column_letter(
                    i) + rowNum] = UserInfo.getUserEfZInfo(nami, userDetails[0])
            case "e/h":
                newWs[get_column_letter(i) + rowNum] = "E" if UserInfo.userTätigkeit(
                    nami, userDetails[0], "hauptamtlicheIDs.csv") == "ERROR: keine Tätigkeit (ERROR)" else "H"
            case "tätigkeit/bula":
                newWs[get_column_letter(i) + rowNum] = UserInfo.userTätigkeit(
                    nami, userDetails[0], "bulaTätigkeitenIDs.csv").split("(")[0]
            case "tätigkeit/bund":
                newWs[get_column_letter(i) + rowNum] = UserInfo.userTätigkeit(
                    nami, userDetails[0], "bundesTätigkeitIDs.csv").split("(")[0]
            case "tätigkeit/bund-365":
                newWs[get_column_letter(i) + rowNum] = UserInfo.userTätigkeit(
                    nami, userDetails[0], "bundesTätigkeitIDs.csv", -365).split("(")[0]
            case "tätigkeit/hauptamtlich":
                newWs[get_column_letter(i) + rowNum] = UserInfo.userTätigkeit(
                    nami, userDetails[0], "hauptamtlicheIDs.csv").split("(")[0]
            case "intakt":
                tempSchulung = UserInfo.userSchulung(
                    nami, userDetails[0], "Bula22 Intakt Schulung")
                newWs[get_column_letter(
                    i) + rowNum] = tempSchulung if tempSchulung != "" and tempSchulung != None else "Keine Intaktschulung!"


def copyFirstLine():
    for i in range(1, sourceWs.max_column + 1):
        newWs[get_column_letter(
            i) + "1"] = sourceWs[get_column_letter(i) + "1"].value


copyFirstLine()
for i in range(2, sourceWs.max_row + 1):
    rowNum = str(i)

    if not sourceWs["B" + rowNum].value:
        newWs["A" + rowNum] = "ERROR: Keine ID"
        continue

    user = sourceWs["B" + rowNum].value
    print(user)
    userDetails = [user, nami.search({"mitgliedsNummber": user}, 10)]
    fillInExcel(userDetails, rowNum)


newWb.save(path + "\\data\\newWBID.xlsx")
sourceWb.close()
newWb.close()
