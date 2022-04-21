import pathlib

from API.APITools import Nami
from API.getUser import UserIDAndData as UID
from decouple import config
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# variables
path = str(pathlib.Path(__file__).parent.resolve())
username = config("USER")
password = config("PASSWORD")


config = []
nami = Nami(config)
nami.auth(username, password)

# excel init
sourceWb = load_workbook(path + "\data\sourceData.xlsx")
sourceWs = sourceWb.active

newWb = Workbook()
newWs = newWb.active


def fillInExcel(userDetails, rowNum):
    if not isinstance(userDetails[0], int):
        newWs["A" + rowNum] = userDetails[0]
        newWs["B" + rowNum] = sourceWs["B" + rowNum].value
        newWs["C" + rowNum] = sourceWs["C" + rowNum].value
        return
    for i in range(1, sourceWs.max_column+1):
        newWs[get_column_letter(i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
        match sourceWs[get_column_letter(i) + "1"].value.lower():
            case "vorname":
                newWs[get_column_letter(i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
            case "nachname":
                newWs[get_column_letter(i) + rowNum] = sourceWs[get_column_letter(i) + rowNum].value
            case "name":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["vorname"] + " " + userDetails[1]["nachname"]
            case "id":
                newWs[get_column_letter(i) + rowNum] = userDetails[0]
            case "straße":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["strasse"]
            case "plz":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["plz"]
            case "plz/ort":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["plz"] + " " + userDetails[1]["ort"]
            case "ort":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["ort"]
            case "adresse":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["strasse"] + " " + userDetails[1]["plz"] + " " + userDetails[1]["ort"]
            case "gechlecht m/w":
                newWs[get_column_letter(i) + rowNum] = "m" if userDetails[1]["geschlecht"] == "männlich" else "w"
            case "geschlecht":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["geschlecht"]
            case "bundesland":
                newWs[get_column_letter(i) + rowNum] = UID.plzZuBundesland(userDetails[1]["plz"])
            case "telefon":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["telefon1"] if userDetails[1]["telefon1"] != "" else userDetails[1]["telefon2"] if userDetails[1]["telefon2"] != "" else userDetails[1]["telefon3"]
            case "telefonnummer":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["telefon1"] if userDetails[1]["telefon1"] != "" else userDetails[1]["telefon2"] if userDetails[1]["telefon2"] != "" else userDetails[1]["telefon3"]
            case "lv":
                newWs[get_column_letter(i) + rowNum] = UID.stammesIdToLV(userDetails[1]["gruppierung"])[1]
            case "geburtstag":
                newWs[get_column_letter(i) + rowNum] = UID.formatDate(str(userDetails[1]["geburtsDatum"])[0:10])
            case "geburtsdatum":
                newWs[get_column_letter(i) + rowNum] = UID.formatDate(str(userDetails[1]["geburtsDatum"])[0:10])
            case "email":
                newWs[get_column_letter(i) + rowNum] = userDetails[1]["email"]
            case "efz":
                newWs[get_column_letter(i) + rowNum] = UID.getUserEfZInfo(nami, userDetails[0])
            case "e/h":
                newWs[get_column_letter(i) + rowNum] = "E" if UID.userTätigkeit(nami, userDetails[0], UID.hauptamtlicheIDs) == "ERROR: keine Tätigkeit (ERROR)" else "H"
            case "tätigkeit/bula":
                newWs[get_column_letter(i) + rowNum] = UID.userTätigkeit(nami, userDetails[0], UID.bulaTätigkeitenIDs).split("(")[0]
            case "tätigkeit/bund":
                newWs[get_column_letter(i) + rowNum] = UID.userTätigkeit(nami, userDetails[0], UID.bundesTaetigkeitIDs).split("(")[0]
            case "tätigkeit/bund-365":
                newWs[get_column_letter(i) + rowNum] = UID.userTätigkeit(nami, userDetails[0], UID.bundesTaetigkeitIDs, -365).split("(")[0]
            case "tätigkeit/hauptamtlich":
                newWs[get_column_letter(i) + rowNum] = UID.userTätigkeit(nami, userDetails[0], UID.hauptamtlicheIDs).split("(")[0]


def copyFirstLine():
    for i in range(1, sourceWs.max_column + 1):
        newWs[get_column_letter(i) + "1"] = sourceWs[get_column_letter(i) + "1"].value


copyFirstLine()
for i in range(2, sourceWs.max_row + 1):
    rowNum = str(i)

    if not sourceWs["B" + rowNum].value or not sourceWs["C" + rowNum].value:
        newWs["A" + rowNum] = "ERROR: Fehlerhafter Name"
        continue

    vornameTemp = sourceWs["B" + rowNum].value.strip().split()[0]
    nachnameTemp = sourceWs["C" + rowNum].value.strip()
    print(vornameTemp, nachnameTemp)
    user = nami.user(vornameTemp, nachnameTemp)
    userDetails = UID.getUserIDAndData(nami, user, vornameTemp, nachnameTemp)
    fillInExcel(userDetails, rowNum)


sourceWb.save(path + "\\data\\sourceData.xlsx")
newWb.save(path + "\\data\\newWB.xlsx")
