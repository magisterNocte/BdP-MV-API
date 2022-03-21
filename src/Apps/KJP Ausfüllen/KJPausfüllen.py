import pathlib

from API.APITools import Nami
from API.getUser import UserIDAndData as UID
from decouple import config
from openpyxl import load_workbook

# variables
path = str(pathlib.Path(__file__).parent.resolve())
username = config("USER")
password = config("PASSWORD")


config = []
nami = Nami(config)
nami.auth(username, password)

# excel init
sourceWb = load_workbook(path + "\data\data.xlsx")  # TODO: evlt eine source und outputdatei einfügen
sourceWs = sourceWb.active


def fillInExcel(userDetails, rowNum):
    if not isinstance(userDetails[0], int):
        sourceWs["C" + rowNum] = userDetails[0]
        return

    sourceWs["C" + rowNum] = userDetails[1]["strasse"]
    sourceWs["D" + rowNum] = userDetails[1]["plz"] + " " + userDetails[1]["ort"]
    sourceWs["E" + rowNum] = "m" if userDetails[1]["geschlecht"] == "männlich" else "w"
    sourceWs["F" + rowNum] = UID.plzZuBundesland(userDetails[1]["plz"])
    sourceWs["G" + rowNum] = "E" if UID.userTätigkeit(nami, userDetails[0], UID.hauptamtlicheIDs) == "ERROR: keine Tätigkeit (ERROR)" else "H"
    sourceWs["H" + rowNum] = UID.userTätigkeit(nami, userDetails[0], UID.hauptamtlicheIDs).split("(")[0]
    sourceWs["I" + rowNum] = UID.stammesIdToLV(userDetails[1]["gruppierung"])[1]
    sourceWs["J" + rowNum] = UID.getUserEfZInfo(nami, userDetails[0])


for i in range(2, sourceWs.max_row + 1):
    rowNum = str(i)

    if not sourceWs["A" + rowNum].value or not sourceWs["B" + rowNum].value:
        sourceWs["C" + rowNum] = "ERROR: Fehlerhafter Name"
        continue

    vornameTemp = sourceWs["A" + rowNum].value.strip().split()[0]
    nachnameTemp = sourceWs["B" + rowNum].value.strip()
    print(vornameTemp, nachnameTemp)
    user = nami.user(vornameTemp, nachnameTemp)
    userDetails = UID.getUserIDAndData(nami, user, vornameTemp, nachnameTemp)
    fillInExcel(userDetails, rowNum)


sourceWb.save(path + "\data\data.xlsx")
