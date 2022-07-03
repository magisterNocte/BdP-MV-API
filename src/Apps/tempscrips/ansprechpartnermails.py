import logging
import pathlib

from decouple import config
from openpyxl import load_workbook
from Tools.APITools import Nami

logging.basicConfig(level=logging.ERROR)
# variables
path = str(pathlib.Path(__file__).parent.resolve())
username = config("MVUSERNAME")
password = config("PASSWORD")


config = []
nami = Nami(config)
nami.auth(username, password)

# excel init
sourceWb = load_workbook(path + "\data\sourceData.xlsx")
sourceWs = sourceWb.active

"""
for i in range(2, sourceWs.max_row + 1):
    if not sourceWs["A" + str(i)].value:
        continue
    print(sourceWs["A" + str(i)].value)
    try:
        sourceWs["E" + str(i)] = ",".join(i["entries_email"] for i in nami.search(
            {"taetigkeitId": [10648], "gruppierung3Id": [int(sourceWs["A" + str(i)].value)]}, 10))
    except:
        sourceWs["E" + str(i)] = "Error"
    try:
        sourceWs["F" + str(i)] = ",".join(i["entries_email"] for i in nami.search(
            {"taetigkeitId": [148, 170], "gruppierung3Id": [int(sourceWs["A" + str(i)].value)]}, 10))
    except:
        sourceWs["F" + str(i)] = "Error"

"""

for i in range(2, sourceWs.max_row + 1):
    if not sourceWs["A" + str(i)].value:
        continue
    print(sourceWs["A" + str(i)].value)

    sourceWs["F" + str(i)] = ",".join(str(i["entries_id"]) for i in nami.search(
        {"taetigkeitId": [148], "gruppierung3Id": [int(sourceWs["A" + str(i)].value)]}, 10))


sourceWb.save(path + "\\data\\sourceData.xlsx")
